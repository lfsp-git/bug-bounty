"""Export findings to various formats: CSV, XLSX, XML."""
import os
import csv
import json
import logging
from datetime import datetime
from typing import List, Dict, Any
from xml.etree import ElementTree as ET

logger = logging.getLogger(__name__)

class ExportFormatter:
    """Formats and exports findings to CSV, XLSX, or XML."""
    
    EXPORT_DIR = "./reports"
    
    def __init__(self):
        os.makedirs(self.EXPORT_DIR, exist_ok=True)
    
    def to_csv(self, findings: List[Dict[str, Any]], filename: str = None) -> str:
        """Export findings to CSV format."""
        if not filename:
            filename = f"findings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        filepath = os.path.join(self.EXPORT_DIR, filename)
        
        try:
            if not findings:
                logger.warning("No findings to export to CSV")
                return filepath
            
            # Extract all unique keys from findings
            keys = set()
            for finding in findings:
                keys.update(finding.keys())
            
            fieldnames = sorted(list(keys))
            
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                for finding in findings:
                    # Convert nested structures to JSON strings for CSV
                    row = {}
                    for key in fieldnames:
                        value = finding.get(key, '')
                        if isinstance(value, (dict, list)):
                            row[key] = json.dumps(value)
                        else:
                            row[key] = str(value)
                    writer.writerow(row)
            
            logger.info(f"Exported {len(findings)} findings to CSV: {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"Failed to export to CSV: {e}")
            return filepath
    
    def to_xlsx(self, findings: List[Dict[str, Any]], filename: str = None) -> str:
        """Export findings to XLSX format (Excel)."""
        if not filename:
            filename = f"findings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = os.path.join(self.EXPORT_DIR, filename)
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return filepath
        
        try:
            if not findings:
                logger.warning("No findings to export to XLSX")
                return filepath
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Findings"
            
            # Extract all unique keys
            keys = set()
            for finding in findings:
                keys.update(finding.keys())
            
            fieldnames = sorted(list(keys))
            
            # Write header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, field in enumerate(fieldnames, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = field
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Write data rows
            for row_num, finding in enumerate(findings, 2):
                for col_num, field in enumerate(fieldnames, 1):
                    value = finding.get(field, '')
                    if isinstance(value, (dict, list)):
                        cell_value = json.dumps(value)
                    else:
                        cell_value = str(value)
                    
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Auto-adjust column widths
            for col_num, field in enumerate(fieldnames, 1):
                max_length = len(str(field))
                for row in ws.iter_rows(min_row=2, max_row=len(findings)+1, min_col=col_num, max_col=col_num):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = adjusted_width
            
            wb.save(filepath)
            logger.info(f"Exported {len(findings)} findings to XLSX: {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"Failed to export to XLSX: {e}")
            return filepath
    
    def to_xml(self, findings: List[Dict[str, Any]], filename: str = None) -> str:
        """Export findings to XML format."""
        if not filename:
            filename = f"findings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xml"
        
        filepath = os.path.join(self.EXPORT_DIR, filename)
        
        try:
            root = ET.Element("findings")
            root.set("count", str(len(findings)))
            root.set("exported", datetime.now().isoformat())
            
            for finding in findings:
                finding_elem = ET.SubElement(root, "finding")
                
                for key, value in finding.items():
                    # Replace invalid XML characters in key
                    safe_key = key.replace(" ", "_").replace("-", "_").lower()
                    if safe_key[0].isdigit():
                        safe_key = "f_" + safe_key
                    
                    elem = ET.SubElement(finding_elem, safe_key)
                    if isinstance(value, (dict, list)):
                        elem.text = json.dumps(value)
                    else:
                        elem.text = str(value)
            
            tree = ET.ElementTree(root)
            tree.write(filepath, encoding='utf-8', xml_declaration=True)
            
            logger.info(f"Exported {len(findings)} findings to XML: {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"Failed to export to XML: {e}")
            return filepath
    
    def export(self, findings: List[Dict[str, Any]], format: str = "csv", filename: str = None) -> str:
        """Export findings in specified format."""
        if format.lower() == "csv":
            return self.to_csv(findings, filename)
        elif format.lower() == "xlsx":
            return self.to_xlsx(findings, filename)
        elif format.lower() == "xml":
            return self.to_xml(findings, filename)
        else:
            logger.error(f"Unknown export format: {format}")
            return ""
