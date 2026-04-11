"""
Hunt3r — Export findings to CSV/XLSX/XML and dry-run preview mode.
"""
from __future__ import annotations

import csv
import glob
import json
import logging
import os
from datetime import datetime
from typing import Any, Dict, List
from xml.etree import ElementTree as ET

logger = logging.getLogger(__name__)

EXPORT_DIR = "reports"


# ---------------------------------------------------------------------------
# Multi-format exporter
# ---------------------------------------------------------------------------
class ExportFormatter:
    """Export a list of finding dicts to CSV, XLSX, or XML."""

    def __init__(self) -> None:
        os.makedirs(EXPORT_DIR, exist_ok=True)

    def _filename(self, ext: str) -> str:
        return os.path.join(EXPORT_DIR, f"findings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}")

    def to_csv(self, findings: List[Dict[str, Any]], filename: str | None = None) -> str:
        filepath = filename or self._filename("csv")
        if not findings:
            logger.warning("No findings to export")
            return filepath
        keys = sorted({k for f in findings for k in f})
        try:
            with open(filepath, "w", newline="", encoding="utf-8") as fh:
                writer = csv.DictWriter(fh, fieldnames=keys)
                writer.writeheader()
                for finding in findings:
                    row = {
                        k: json.dumps(v) if isinstance(v, (dict, list)) else str(v)
                        for k, v in finding.items()
                    }
                    writer.writerow(row)
            logger.info(f"Exported {len(findings)} findings → {filepath}")
        except OSError as e:
            logger.error(f"CSV export failed: {e}")
        return filepath

    def to_xlsx(self, findings: List[Dict[str, Any]], filename: str | None = None) -> str:
        filepath = filename or self._filename("xlsx")
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return filepath
        if not findings:
            return filepath
        keys = sorted({k for f in findings for k in f})
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Findings"
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col, key in enumerate(keys, 1):
            cell = ws.cell(row=1, column=col, value=key)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for row, finding in enumerate(findings, 2):
            for col, key in enumerate(keys, 1):
                val = finding.get(key, "")
                cell = ws.cell(
                    row=row,
                    column=col,
                    value=json.dumps(val) if isinstance(val, (dict, list)) else str(val),
                )
                cell.alignment = Alignment(wrap_text=True)
        for col, key in enumerate(keys, 1):
            max_len = max(
                (len(str(ws.cell(r, col).value or "")) for r in range(1, len(findings) + 2)),
                default=len(key),
            )
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 2, 50)
        try:
            wb.save(filepath)
        except OSError as e:
            logger.error(f"XLSX export failed: {e}")
        return filepath

    def to_xml(self, findings: List[Dict[str, Any]], filename: str | None = None) -> str:
        filepath = filename or self._filename("xml")
        root = ET.Element("findings", count=str(len(findings)), exported=datetime.now().isoformat())
        for finding in findings:
            fe = ET.SubElement(root, "finding")
            for key, value in finding.items():
                safe = re.sub(r"[^a-z0-9_]", "_", key.lower())
                if safe and safe[0].isdigit():
                    safe = "f_" + safe
                if not safe:
                    continue
                ET.SubElement(fe, safe).text = (
                    json.dumps(value) if isinstance(value, (dict, list)) else str(value)
                )
        try:
            ET.ElementTree(root).write(filepath, encoding="utf-8", xml_declaration=True)
        except OSError as e:
            logger.error(f"XML export failed: {e}")
        return filepath

    def to_pdf(self, findings: List[Dict[str, Any]], filename: str | None = None) -> str:
        """Export findings as a print-ready HTML file (save as PDF from browser).

        Attempts to use fpdf2 if installed for true PDF output.
        Falls back to styled HTML which can be printed to PDF via browser.
        """
        # Try true PDF with fpdf2 (optional dep)
        try:
            from fpdf import FPDF
            return self._to_pdf_fpdf(findings, filename)
        except ImportError:
            pass

        # Fallback: styled HTML that renders identically when printed/saved as PDF
        filepath = filename or self._filename("pdf.html")
        sev_colors = {
            "critical": "#d32f2f", "high": "#f57c00",
            "medium": "#fbc02d", "low": "#388e3c", "info": "#1976d2",
        }
        rows = []
        for f in findings:
            sev = (f.get("severity") or f.get("info", {}).get("severity", "info")).lower()
            color = sev_colors.get(sev, "#555")
            tid = f.get("template-id", "unknown")
            matched = f.get("matched-at", "")
            cve = f.get("info", {}).get("classification", {}).get("cve-id", [""])[0] if isinstance(
                f.get("info", {}).get("classification", {}).get("cve-id"), list) else ""
            rows.append(
                f'<tr><td style="color:{color};font-weight:bold">{sev.upper()}</td>'
                f'<td>{tid}</td><td style="word-break:break-all">{matched}</td>'
                f'<td>{cve}</td></tr>'
            )
        table_body = "\n".join(rows) if rows else '<tr><td colspan="4">No findings</td></tr>'
        html = f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8">
<title>Hunt3r Findings Report</title>
<style>
  body{{font-family:Arial,sans-serif;font-size:12px;margin:20px}}
  h1{{color:#1a237e}} table{{width:100%;border-collapse:collapse}}
  th{{background:#1a237e;color:#fff;padding:6px 10px;text-align:left}}
  td{{padding:5px 10px;border-bottom:1px solid #ddd}}
  tr:nth-child(even){{background:#f5f5f5}}
  @media print{{body{{margin:0}} .no-print{{display:none}}}}
</style></head>
<body>
<h1>🎯 Hunt3r Security Findings</h1>
<p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M UTC')} &nbsp;|&nbsp; Total findings: {len(findings)}</p>
<p class="no-print"><em>Tip: Use browser Print → Save as PDF for a PDF copy.</em></p>
<table>
  <thead><tr><th>Severity</th><th>Template</th><th>Matched At</th><th>CVE</th></tr></thead>
  <tbody>{table_body}</tbody>
</table>
</body></html>"""
        try:
            with open(filepath, "w", encoding="utf-8") as fh:
                fh.write(html)
            logger.info(f"PDF report (HTML) saved: {filepath} — open in browser and print to PDF")
        except OSError as e:
            logger.error(f"PDF export failed: {e}")
        return filepath

    def _to_pdf_fpdf(self, findings: List[Dict[str, Any]], filename: str | None = None) -> str:
        """True PDF export using fpdf2 (if installed)."""
        from fpdf import FPDF
        filepath = filename or self._filename("pdf")
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, "Hunt3r Security Findings", ln=True)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M UTC')} | Total: {len(findings)}", ln=True)
        pdf.ln(3)
        col_w = [22, 55, 85, 28]
        headers = ["Severity", "Template ID", "Matched At", "CVE"]
        pdf.set_fill_color(26, 35, 126)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 8)
        for h, w in zip(headers, col_w):
            pdf.cell(w, 7, h, border=1, fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 7)
        sev_rgb = {
            "critical": (211, 47, 47), "high": (245, 124, 0),
            "medium": (251, 192, 45), "low": (56, 142, 60),
        }
        for f in findings:
            sev = (f.get("severity") or f.get("info", {}).get("severity", "info")).lower()
            r, g, b = sev_rgb.get(sev, (85, 85, 85))
            pdf.set_text_color(r, g, b)
            tid = str(f.get("template-id", ""))[:40]
            matched = str(f.get("matched-at", ""))[:60]
            cve_list = f.get("info", {}).get("classification", {}).get("cve-id", [])
            cve = cve_list[0] if isinstance(cve_list, list) and cve_list else ""
            row = [sev.upper(), tid, matched, cve[:20]]
            for val, w in zip(row, col_w):
                pdf.cell(w, 5, val, border=1)
            pdf.ln()
            pdf.set_text_color(0, 0, 0)
        try:
            pdf.output(filepath)
            logger.info(f"PDF saved: {filepath}")
        except Exception as e:
            logger.error(f"PDF fpdf2 export failed: {e}")
        return filepath

    def export(self, findings: List[Dict[str, Any]], fmt: str = "csv") -> str:
        """Dispatch to the appropriate format handler."""
        handlers = {
            "csv": self.to_csv,
            "xlsx": self.to_xlsx,
            "xml": self.to_xml,
            "pdf": self.to_pdf,
        }
        handler = handlers.get(fmt.lower())
        if not handler:
            logger.error(f"Unknown export format: {fmt}")
            return ""
        return handler(findings)


# ---------------------------------------------------------------------------
# Dry-run preview
# ---------------------------------------------------------------------------
def run_dry_run() -> None:
    """Preview scan targets without executing any tools."""
    from core.ui import ui_log, Colors  # deferred import
    from recon.platforms import PlatformManager

    try:
        pm = PlatformManager()
        all_targets: List[Dict] = []

        for platform in ("h1", "it"):
            try:
                targets = pm.get_all_programs_from_platform(platform)
                all_targets.extend(targets)
                ui_log(f"DRY-RUN:{platform.upper()}", f"{len(targets)} targets loaded", Colors.SUCCESS)
            except Exception as e:
                ui_log(f"DRY-RUN:{platform.upper()}", f"Failed: {str(e)[:50]}", Colors.WARNING)

        if not all_targets:
            ui_log("DRY-RUN", "No targets found across all platforms", Colors.ERROR)
            return

        print("\n" + "=" * 60)
        print(f"DRY RUN  —  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 60)
        print(f"Total targets : {len(all_targets)}")
        print("=" * 60)
        for t in all_targets[:30]:
            domains = ", ".join((t.get("domains") or [])[:3])
            ui_log("TARGET", f"{t.get('handle', '?')} — {domains}", Colors.DIM)
        if len(all_targets) > 30:
            print(f"  ...and {len(all_targets) - 30} more")

        # Save JSON report
        os.makedirs(EXPORT_DIR, exist_ok=True)
        report_file = os.path.join(EXPORT_DIR, f"dry_run_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        with open(report_file, "w", encoding="utf-8") as f:
            json.dump(
                {
                    "timestamp": datetime.now().isoformat(),
                    "mode": "dry-run",
                    "total": len(all_targets),
                    "targets": [
                        {"handle": t.get("handle"), "domains": t.get("domains", [])}
                        for t in all_targets
                    ],
                },
                f,
                indent=2,
            )
        ui_log("DRY-RUN", f"Report: {report_file}", Colors.SUCCESS)
    except Exception as e:
        logger.error(f"Dry run failed: {e}", exc_info=True)
        from core.ui import ui_log, Colors
        ui_log("DRY-RUN", f"Error: {str(e)[:60]}", Colors.ERROR)


# needed by to_xml
import re
